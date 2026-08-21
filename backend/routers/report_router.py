import io
import json
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
import models
from deps import get_current_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/pdf")
def skin_report_pdf(user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    from fpdf import FPDF

    profile = db.query(models.SkinProfile).filter(models.SkinProfile.user_id == user.id).first()
    latest_score = db.query(models.SkinHealthScore).filter(
        models.SkinHealthScore.user_id == user.id
    ).order_by(models.SkinHealthScore.created_at.desc()).first()
    latest_assessment = db.query(models.SkinAssessment).filter(
        models.SkinAssessment.user_id == user.id
    ).order_by(models.SkinAssessment.created_at.desc()).first()
    routines = db.query(models.Routine).filter(
        models.Routine.user_id == user.id, models.Routine.is_active == True
    ).all()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Skin Health Report", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Name: {user.full_name or user.email}", ln=True)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Skin Profile", ln=True)
    pdf.set_font("Helvetica", "", 11)
    if profile:
        pdf.multi_cell(0, 7, f"Skin type: {profile.skin_type or 'N/A'}\n"
                              f"Age group: {profile.age_group or 'N/A'}\n"
                              f"Concerns: {', '.join(json.loads(profile.concerns_json or '[]')) or 'None'}\n"
                              f"Allergies: {', '.join(json.loads(profile.allergies_json or '[]')) or 'None'}")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Skin Health Score", ln=True)
    pdf.set_font("Helvetica", "", 11)
    if latest_score:
        pdf.multi_cell(0, 7,
            f"Overall Score: {latest_score.overall_score}/100\n"
            f"  - Condition (35%): {latest_score.condition_score}\n"
            f"  - Lifestyle (20%): {latest_score.lifestyle_score}\n"
            f"  - Sleep (15%): {latest_score.sleep_score}\n"
            f"  - Routine Consistency (20%): {latest_score.routine_score}\n"
            f"  - Hydration (10%): {latest_score.hydration_score}")
    else:
        pdf.cell(0, 7, "No assessment run yet.", ln=True)
    pdf.ln(4)

    if latest_assessment:
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Risk Flags", ln=True)
        pdf.set_font("Helvetica", "", 11)
        flags = json.loads(latest_assessment.risk_flags_json or "[]")
        pdf.multi_cell(0, 7, "\n".join(f"- {f}" for f in flags) or "None")
        pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Active Routine", ln=True)
    pdf.set_font("Helvetica", "", 11)
    for r in routines:
        steps = json.loads(r.steps_json)
        pdf.cell(0, 7, f"{r.period.capitalize()}:", ln=True)
        for s in steps:
            pdf.cell(0, 6, f"  {s['order']}. {s['step']} - {s.get('suggested_ingredient') or s['category']}", ln=True)

    pdf_bytes = bytes(pdf.output())
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=skin_health_report.pdf"},
    )


@router.get("/excel")
def progress_report_excel(user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    from openpyxl import Workbook

    scores = db.query(models.SkinHealthScore).filter(
        models.SkinHealthScore.user_id == user.id
    ).order_by(models.SkinHealthScore.created_at.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Skin Health Scores"
    ws.append(["Date", "Overall", "Condition", "Lifestyle", "Sleep", "Routine", "Hydration"])
    for s in scores:
        ws.append([
            s.created_at.strftime("%Y-%m-%d %H:%M"), s.overall_score, s.condition_score,
            s.lifestyle_score, s.sleep_score, s.routine_score, s.hydration_score,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=skin_progress_report.xlsx"},
    )
